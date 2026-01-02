from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from utils.db import (
    conectar_bd,
    obtener_nombre_empleado,
    obtener_datos_cliente,
    obtener_descripcion_producto
)
import os
import subprocess
import sys

def convertir_a_pdf(ruta_excel, carpeta_salida):
    print("--- 1. Iniciando conversión a PDF ---")
    
    if sys.platform == "win32":
        ejecutable = r"C:\Program Files\LibreOffice\program\soffice.exe"
        print(f"--- 2. OS detectado: Windows. Usando ruta: {ejecutable}")
        if not os.path.exists(ejecutable):
            print("--- ERROR FATAL: La ruta al ejecutable de LibreOffice en Windows NO EXISTE. Verifica la ruta. ---")
            return None
    else:
        ejecutable = "libreoffice"
        print(f"--- 2. OS detectado: Linux/Otro. Usando comando: {ejecutable}")

    comando = [
        ejecutable,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", carpeta_salida,
        ruta_excel
    ]
    print(f"--- 3. Ejecutando comando: {' '.join(comando)}")
    try:
        subprocess.run(comando, check=True, timeout=30)
        print("--- 4. Conversión a PDF exitosa ---")
        return os.path.splitext(os.path.basename(ruta_excel))[0] + ".pdf"
    except FileNotFoundError:
        print("--- ERROR FATAL: 'FileNotFoundError'. El sistema no pudo encontrar el comando. ---")
        return None
    except Exception as e:
        print(f"--- ERROR INESPERADO durante la conversión a PDF: {e} ---")
        return None

def generar_nota_pedido(cotizacion_id):
    print(f"\n--- INICIANDO GENERACIÓN DE NOTA DE PEDIDO PARA ID: {cotizacion_id} ---")
    conn = None
    try:
        conn = conectar_bd()
        cur = conn.cursor()
        
        cur.execute("SELECT cliente_id, creado_por FROM cotizaciones WHERE id_cotizacion = %s", (cotizacion_id,))
        fila = cur.fetchone()
        if not fila:
            raise ValueError("Cotización no encontrada")
        cliente_id, matricula = fila
        
        cur.execute("SELECT id_producto, cantidad, precio_unitario, total FROM detalle_cotizacion WHERE id_cotizacion = %s ORDER BY id_producto", (cotizacion_id,))
        productos = cur.fetchall()
        
        base_dir = os.path.dirname(__file__)
        plantilla_path = os.path.join(base_dir, 'plantilla_nota_entrega.xlsx')
        wb = load_workbook(plantilla_path)
        ws = wb.active
        
        hoy = datetime.today()
        nombre_empleado = obtener_nombre_empleado(matricula)
        cliente = obtener_datos_cliente(cliente_id)
        
        ws['F3'] = hoy.strftime('%d/%m/%Y')
        ws['F4'] = f"{cotizacion_id:05d}"
        ws['F5'] = cliente_id
        ws['A10'] = f"Nombre: {cliente[0]}"
        direccion = f"Dirección: {cliente[1] or ''} {cliente[2] or ''}".strip()
        if cliente[3]: direccion += f", Int. {cliente[3]}"
        ws['A11'] = direccion
        ws['A12'] = f"{cliente[4] or ''}, {cliente[5] or ''}, {cliente[6] or ''}, {cliente[7] or ''}"
        ws['A13'] = f"Teléfono: {cliente[8] or ''}"
        ws['A14'] = f"Correo: {cliente[9] or ''}"
        ws['A7'] = f"Entregado por: {nombre_empleado}"
        ws['A15'] = f"RFC: {cliente[10] or ''}"
        
        fila_inicio = 18
        for i, (id_producto, cantidad, precio_unitario, total) in enumerate(productos):
            fila = fila_inicio + i
            descripcion = obtener_descripcion_producto(id_producto)
            ws[f"A{fila}"] = descripcion
            ws[f"C{fila}"] = float(precio_unitario)
            ws[f"D{fila}"] = int(cantidad)
            ws[f"F{fila}"] = float(total)
            
        logo_path = os.path.join(base_dir, 'static', 'logo.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 60
            img.height = 60
            ws.add_image(img, "A1")
            
        carpeta_salida = os.path.normpath(os.path.join(base_dir, '..', 'notas_pedido'))
        os.makedirs(carpeta_salida, exist_ok=True)
        ruta_excel = os.path.join(carpeta_salida, f"nota_pedido_{cliente_id}_{cotizacion_id:05d}.xlsx")
        print(f"--- Guardando archivo Excel en: {ruta_excel} ---")
        wb.save(ruta_excel)
        
        convertir_a_pdf(ruta_excel, carpeta_salida)
        return ruta_excel
    finally:
        if conn:
            cur.close()
            conn.close()
        print("--- FIN DE GENERACIÓN DE NOTA DE PEDIDO ---")