# generar_cotizacion.py
from flask import Blueprint, request, redirect, url_for, session, jsonify, render_template
from datetime import date, timedelta, datetime
from utils.db import conectar_bd
import os
from flask import send_from_directory, abort
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import subprocess
from utils.db import obtener_nombre_empleado
import sys


# Blueprint con el nombre que usan tus plantillas (cotizaciones)
generar_cotizacion_bp = Blueprint("cotizaciones", __name__, url_prefix="/cotizaciones")

# ─────────────────────────────
# Normalización de productos
# ─────────────────────────────
# --- MODIFICACIÓN: Añadir SKUs 1kg y 3kg ---
SKU_1KG  = "BR1KG"
SKU_3KG  = "BR3KG"
SKU_5KG  = "BR5KG"
SKU_15KG = "BR15KG"

ALIAS_A_SKU = {
    # Aliases 1kg
    "1": SKU_1KG, "1KG": SKU_1KG, "B1KG": SKU_1KG, "BOLSA 1": SKU_1KG, "BOLSA 1KG": SKU_1KG,
    "1 KILOGRAMO": SKU_1KG, "BOLSA DE HIELO DE 1 KILOGRAMO": SKU_1KG,

    # Aliases 3kg
    "3": SKU_3KG, "3KG": SKU_3KG, "B3KG": SKU_3KG, "BOLSA 3": SKU_3KG, "BOLSA 3KG": SKU_3KG,
    "3 KILOGRAMOS": SKU_3KG, "BOLSA DE HIELO DE 3 KILOGRAMOS": SKU_3KG,

    # Aliases 5kg
    "5": SKU_5KG, "5KG": SKU_5KG, "B5KG": SKU_5KG, "BOLSA 5": SKU_5KG, "BOLSA 5KG": SKU_5KG,
    "5 KILOGRAMOS": SKU_5KG, "BOLSA DE HIELO DE 5 KILOGRAMOS": SKU_5KG,

    # Aliases 15kg
    "15": SKU_15KG, "15KG": SKU_15KG, "B15KG": SKU_15KG, "BOLSA 15": SKU_15KG, "BOLSA 15KG": SKU_15KG,
    "15 KILOGRAMOS": SKU_15KG, "BOLSA DE HIELO DE 15 KILOGRAMOS": SKU_15KG,
}
# --- MODIFICACIÓN TERMINA ---


def normalizar_id_producto(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    return ALIAS_A_SKU.get(s, s)

def existe_producto(cur, id_producto: str) -> bool:
    cur.execute("SELECT 1 FROM productos WHERE id_producto = %s", (id_producto,))
    return cur.fetchone() is not None

# ... (El resto del archivo 'generar_cotizacion.py' (funciones registrar_cotizacion,
# descargar_cotizacion, generar_web, convertir_a_pdf, generar_archivo_cotizacion)
# no necesita más modificaciones, ya que la lógica de normalización actualizada
# y la carga de productos desde la BD (que ya contiene 1kg y 3kg)
# son suficientes para manejar los nuevos tamaños.)

# ─────────────────────────────
# Núcleo: registrar cotización
# ─────────────────────────────
# En app/generar_cotizacion.py

def registrar_cotizacion(cliente_id: str, matricula: str, productos: list, *,
                         fecha: date | None = None,
                         dias_validez: int = 7,
                         tipo_inicial: str = "Cotización",
                         estatus_inicial: str = "Nuevo") -> int:
    """
    Inserta en 'cotizaciones' y 'detalle_cotizacion'.
    productos: lista de dicts o tuplas con:
      id_producto, descripcion, cantidad, precio_unitario, total
    """
    fecha = fecha or date.today()
    valido_hasta = fecha + timedelta(days=dias_validez)

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Cabecera (esto ya funciona bien)
        cur.execute("""
            INSERT INTO cotizaciones
                (fecha, valido_hasta, cliente_id, atendido_por, notas,
                 fecha_hora_creacion, creado_por, tipo, estatus)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id_cotizacion
        """, (
            fecha, valido_hasta, cliente_id,
            matricula,              # atendido_por
            None,                   # notas
            datetime.now(),         # fecha_hora_creacion
            matricula,              # creado_por
            tipo_inicial, estatus_inicial
        ))
        id_cotizacion = cur.fetchone()[0]

        # --- INICIA LA SECCIÓN CORREGIDA ---
        # Detalle de la cotización
        for item in productos:
            # 1. Obtener datos del producto enviado desde el formulario
            raw_id_producto = item.get("id_producto")
            cantidad = int(item.get("cantidad", 0) or 0)
            precio_unitario = float(item.get("precio_unitario", 0) or 0.0)
            total = float(item.get("total", cantidad * precio_unitario))

            # 2. Normalizar y validar el ID del producto
            id_producto = normalizar_id_producto(raw_id_producto)
            if not id_producto:
                raise ValueError("Falta id_producto en un renglón del detalle.")
            if not existe_producto(cur, id_producto):
                raise ValueError(
                    f"El producto '{raw_id_producto}' no existe (normalizado a '{id_producto}'). "
                    f"Crea el SKU en la tabla productos o corrige el ID."
                )

            # 3. ¡IMPORTANTE! Buscar la descripción correcta en la base de datos
            cur.execute("SELECT descripcion FROM public.productos WHERE id_producto = %s", (id_producto,))
            resultado_producto = cur.fetchone()
            descripcion_correcta = resultado_producto[0] if resultado_producto else "Descripción no encontrada"

            # 4. Insertar en la base de datos con TODOS los datos correctos
            cur.execute("""
                INSERT INTO detalle_cotizacion
                    (id_cotizacion, id_producto, descripcion, cantidad, precio_unitario, total)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
            """, (id_cotizacion, id_producto, descripcion_correcta, cantidad, precio_unitario, total))
        # --- FIN DE LA SECCIÓN CORREGIDA ---

        conn.commit()
        return id_cotizacion

    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
# ─────────────────────────────
# Descarga de archivos generados
# ─────────────────────────────
@generar_cotizacion_bp.route("/descargar/<path:filename>")
def descargar_cotizacion(filename):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    posibles_dirs = [
        os.path.normpath(os.path.join(base_dir, "..", "cotizaciones")),
        os.path.normpath(os.path.join(base_dir, "..", "notas_pedido")),
        os.path.normpath(os.path.join(base_dir, "cotizaciones"))
    ]
    for carpeta in posibles_dirs:
        ruta = os.path.join(carpeta, filename)
        if os.path.isfile(ruta):
            return send_from_directory(carpeta, filename, as_attachment=True)
    abort(404, description=f"Archivo no encontrado: {filename}")

# ─────────────────────────────
# UI + API en el MISMO endpoint
# ─────────────────────────────
# En app/generar_cotizacion.py
    
@generar_cotizacion_bp.route("/generar", methods=["GET", "POST"])
def generar_web():
    conn = None  # Definimos conn aquí para que esté disponible en el finally
    
    # --- Función auxiliar para cargar datos ---
    # Esto evita repetir código. Se llamará siempre que se renderice la página.
    def render_form_with_data(error_message=None):
        nonlocal conn
        try:
            if not conn or conn.closed:
                conn = conectar_bd()
            cur = conn.cursor()
            
            # Consultar clientes
            cur.execute("SELECT id_cliente, nombre FROM public.clientes ORDER BY id_cliente")
            clientes = cur.fetchall()
            
            # Consultar productos
            cur.execute("SELECT id_producto, nombre, descripcion, precio_sugerido FROM public.productos ORDER BY id_producto")
            productos_bd = cur.fetchall()

            return render_template("generar_cotizacion.html", 
                                   clientes=clientes, 
                                   productos_bd=productos_bd, 
                                   error=error_message)
        except Exception as e:
            # Si hay un error cargando datos, lo mostramos
            return render_template("generar_cotizacion.html", 
                                   clientes=[], 
                                   productos_bd=[], 
                                   error=f"Error al cargar datos: {e}")

    try:
        # --- Lógica para POST (cuando se envía el formulario) ---
        if request.method == "POST":
            cliente_id = (request.form.get("cliente_id") or "").strip()
            matricula = (session.get("usuario") or "").strip()
            ids = request.form.getlist("id_producto[]")
            cants = request.form.getlist("cantidad[]")
            precios = request.form.getlist("precio_unitario[]")

            # Validaciones
            if not cliente_id:
                return render_form_with_data(error_message="El ID del cliente es requerido.")
            if not matricula:
                return render_form_with_data(error_message="No se pudo identificar al usuario. Inicie sesión de nuevo.")
            
            productos = []
            for i in range(len(ids)):
                if ids[i].strip():
                    cantidad = int(cants[i] or 0)
                    pu = float(precios[i] or 0)
                    productos.append({
                        "id_producto": ids[i], "descripcion": "",
                        "cantidad": cantidad, "precio_unitario": pu,
                        "total": cantidad * pu
                    })

            if not productos:
                return render_form_with_data(error_message="Debe agregar al menos un producto.")

            # Si todo es válido, registrar y redirigir
            id_cot = registrar_cotizacion(cliente_id, matricula, productos)
            generar_archivo_cotizacion(id_cot) 
            return redirect(url_for("ges_cotizaciones.detalle_cotizacion", id_cotizacion=id_cot))

        # --- Lógica para GET (cuando se carga la página por primera vez) ---
        return render_form_with_data()

    except Exception as e:
        # Error general durante el POST
        return render_form_with_data(error_message=f"Error inesperado: {e}")
    finally:
        # Asegurarse de que la conexión se cierre
        if conn and not conn.closed:
            conn.close()

# === PEGA ESTAS DOS FUNCIONES AL FINAL DE app/generar_cotizacion.py ===

# Al final de app/generar_cotizacion.py

def convertir_a_pdf(ruta_excel, carpeta_salida):
    # --- LÓGICA PARA DETECTAR EL SISTEMA OPERATIVO ---
    if sys.platform == "win32":
        # Si es Windows, usa la ruta completa
        ejecutable = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else:
        # Si es Linux (o cualquier otro), usa el comando simple
        ejecutable = "libreoffice"
    # ----------------------------------------------------

    comando = [
        ejecutable, "--headless", "--convert-to", "pdf",
        "--outdir", carpeta_salida, ruta_excel
    ]
    try:
        subprocess.run(comando, check=True, capture_output=True, text=True)
        return True
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        print(f"❌ Error al convertir a PDF. Asegúrate de que LibreOffice esté instalado y en el PATH (para Linux) o que la ruta sea correcta (para Windows). Error: {e}")
        return False

def generar_archivo_cotizacion(cotizacion_id):
    """Toma una cotización de la BD y genera su archivo XLSX y PDF."""
    conn = None
    try:
        conn = conectar_bd()
        cur = conn.cursor()

        cur.execute("""
            SELECT c.fecha, c.valido_hasta, c.cliente_id, c.creado_por, cl.nombre, 
                   cl.calle, cl.numero_exterior, cl.numero_interior, cl.colonia, cl.codigo_postal, 
                   cl.municipio, cl.estado, cl.telefono, cl.correo, cl.rfc
            FROM cotizaciones c
            JOIN clientes cl ON c.cliente_id = cl.id_cliente
            WHERE c.id_cotizacion = %s
        """, (cotizacion_id,))
        cot_data = cur.fetchone()
        if not cot_data:
            raise ValueError("No se encontró la cotización para generar el archivo.")

        (fecha, valido_hasta, cliente_id, matricula, nombre_cliente, calle, num_ext, num_int, 
         colonia, cp, municipio, estado, telefono, correo, rfc_cliente) = cot_data

        cur.execute("SELECT descripcion, cantidad, precio_unitario, total FROM detalle_cotizacion WHERE id_cotizacion = %s", (cotizacion_id,))
        productos = cur.fetchall()

        base_dir = os.path.dirname(__file__)
        plantilla_path = os.path.join(base_dir, 'plantilla_cotizacion.xlsx')
        wb = load_workbook(plantilla_path)
        ws = wb.active

        ws['F3'] = fecha.strftime('%d/%m/%Y')
        ws['F4'] = f"{cotizacion_id:05d}"
        ws['F5'] = cliente_id
        ws['F6'] = valido_hasta.strftime('%d/%m/%Y')
        ws['A7'] = f"Le atiende: {obtener_nombre_empleado(matricula)}"
        ws['A10'] = nombre_cliente
        # ... (Llenado de datos del cliente y productos como en la versión anterior) ...

        fila_inicio = 18
        subtotal = 0
        for i, (descripcion, cantidad, precio_unitario, total) in enumerate(productos):
            fila = fila_inicio + i
            ws[f'A{fila}'] = descripcion
            ws[f'C{fila}'] = float(precio_unitario)
            ws[f'D{fila}'] = int(cantidad)
            ws[f'F{fila}'] = float(total)
            subtotal += float(total)

        ws['F30'] = subtotal
        ws['F35'] = subtotal

        # --- Insertar Logo ---
        logo_path = os.path.join(base_dir, 'static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"No se pudo agregar el logo a la cotización: {e}")

        carpeta_salida = os.path.normpath(os.path.join(base_dir, '..', 'cotizaciones'))
        os.makedirs(carpeta_salida, exist_ok=True)
        nombre_archivo = f"cotizacion_{cliente_id}_{cotizacion_id:05d}.xlsx"
        ruta_excel = os.path.join(carpeta_salida, nombre_archivo)
        wb.save(ruta_excel)

        convertir_a_pdf(ruta_excel, carpeta_salida)
        return ruta_excel
    finally:
        if conn:
            cur.close()
            conn.close()