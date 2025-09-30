# generar_nota_pedido.py
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from utils.db import (
    conectar_bd,
    obtener_nombre_empleado,
    obtener_datos_cliente,
    obtener_descripcion_producto
)
import os
import subprocess

def generar_nota_pedido(cotizacion_id):
    # --- Datos base
    conn = conectar_bd()
    cur = conn.cursor()

    cur.execute("""
        SELECT cliente_id, creado_por
        FROM cotizaciones
        WHERE id_cotizacion = %s
    """, (cotizacion_id,))
    fila = cur.fetchone()
    if not fila:
        cur.close(); conn.close()
        raise ValueError("Cotización no encontrada")

    cliente_id, matricula = fila

    cur.execute("""
        SELECT id_producto, cantidad, precio_unitario, total
        FROM detalle_cotizacion
        WHERE id_cotizacion = %s
        ORDER BY id_producto
    """, (cotizacion_id,))
    productos = cur.fetchall()

    cur.close()
    conn.close()

    # --- Plantilla
    plantilla_path = os.path.join(os.path.dirname(__file__), 'plantilla_nota_entrega.xlsx')
    wb = load_workbook(plantilla_path)
    # Ajusta el nombre de hoja si difiere en tu archivo
    ws = wb.active if "NOTA_ENTREGA" not in wb.sheetnames else wb["NOTA_ENTREGA"]

    hoy = datetime.today()
    nombre_empleado = obtener_nombre_empleado(matricula)
    cliente = obtener_datos_cliente(cliente_id)  # [nombre, calle, ext, int, col, mun, edo, cp, tel, correo]

    # --- Encabezados
    ws['F3'] = hoy.strftime('%d/%m/%Y')
    ws['F4'] = f"{cotizacion_id:05d}"  # No. de pedido
    ws['F5'] = cliente_id

    ws['A10'] = f"Nombre: {cliente[0]}"
    direccion = f"Dirección: {cliente[1]} {cliente[2]}".strip()
    if cliente[3]:
        direccion += f", Int. {cliente[3]}"
    ws['A11'] = direccion
    ws['A12'] = f"{cliente[4]}, {cliente[5]}, {cliente[6]}, {cliente[7]}"
    ws['A13'] = f"Teléfono: {cliente[8]}"
    ws['A14'] = f"Correo: {cliente[9]}"
    ws['A7']  = f"Entregado por: {nombre_empleado}"
    ws['A15'] = f"RFC: {cliente[10]}"
    

    # --- Items
    fila_inicio = 18  
    for i, (id_producto, cantidad, precio_unitario, total) in enumerate(productos):
        fila = fila_inicio + i
        descripcion = obtener_descripcion_producto(id_producto)
        ws[f"A{fila}"] = descripcion
        ws[f"C{fila}"] = float(precio_unitario)
        ws[f"D{fila}"] = int(cantidad)
        ws[f"F{fila}"] = float(total)

    # --- Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 60; img.height = 60
            ws.add_image(img, "A1")
        except Exception:
            pass  # si la hoja está protegida u otra cosa, no rompas

    # --- Salida
    carpeta_salida = os.path.join(os.path.dirname(__file__), '../notas_pedido')
    os.makedirs(carpeta_salida, exist_ok=True)
    ruta_excel = os.path.join(carpeta_salida, f"nota_pedido_{cliente_id}_{cotizacion_id:05d}.xlsx")
    wb.save(ruta_excel)

    convertir_a_pdf(ruta_excel, carpeta_salida)
    return ruta_excel

def convertir_a_pdf(ruta_excel, carpeta_salida):
    # Usa tu LibreOffice headless (misma forma que cotización)
    comando = [
        "libreoffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", carpeta_salida,
        ruta_excel
    ]
    try:
        subprocess.run(comando, check=True)
        return os.path.splitext(os.path.basename(ruta_excel))[0] + ".pdf"
    except subprocess.CalledProcessError:
        print("❌ Error al convertir a PDF.")
        return None
